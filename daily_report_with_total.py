import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 🔧 Paths
folder_path = "employee_data"
master_file = "master_report.xlsx"

# 📆 Aaj ki date
today = datetime.today().strftime("%Y-%m-%d")

# 🔁 Daily data collect
daily_data = []
total_self = total_cash = total_onl = total_tran = 0

for file in os.listdir(folder_path):
    if file.endswith(".xlsx") and not file.startswith("~$"):
        path = os.path.join(folder_path, file)
        df = pd.read_excel(path)

        # Ignore summary rows
        df = df[~df['WORK'].astype(str).str.upper().isin(['TOTAL', 'OVERALL'])]

        s = df['SELF'].sum()
        c = df['CASH'].sum()
        o = df['ONL'].sum()
        t = df['TRAN'].sum()

        daily_data.append([file.replace('.xlsx', ''), s, c, o, t])

        total_self += s
        total_cash += c
        total_onl += o
        total_tran += t

# ➕ Add TOTAL row
daily_data.append(['TOTAL', total_self, total_cash, total_onl, total_tran])

# 🧾 Create today's DataFrame
df_today = pd.DataFrame(daily_data, columns=['EMPLOYEE', 'SELF', 'CASH', 'ONL', 'TRAN'])

# ➕ Insert DATE column only for first row
df_today.insert(0, 'DATE', '')
df_today.at[0, 'DATE'] = today

# ➕ Add 2 empty rows after today’s block
empty = pd.DataFrame([['', '', '', '', '', ''], ['', '', '', '', '', '']], columns=df_today.columns)

# 📁 Merge with existing file or create new
if os.path.exists(master_file):
    existing_df = pd.read_excel(master_file)
    final_df = pd.concat([existing_df, df_today, empty], ignore_index=True)
else:
    final_df = pd.concat([df_today, empty], ignore_index=True)

# 💾 Save updated data
final_df.to_excel(master_file, index=False)

# 🎨 Styling part: Highlight TOTAL row + format
wb = load_workbook(master_file)
ws = wb.active

# Format styles
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
bold_font = Font(bold=True)
center_align = Alignment(horizontal='center', vertical='center')

# Auto-adjust column width
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Apply highlight to TOTAL rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    emp_cell = row[1]  # EMPLOYEE column (B)
    if emp_cell.value and str(emp_cell.value).strip().upper() == "TOTAL":
        for cell in row:
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.alignment = center_align

# Save with style
wb.save(master_file)

print("✅ DONE: Report updated with data, TOTAL highlight, formatting, spacing.")
